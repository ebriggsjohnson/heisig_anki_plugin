"""Build Anki-importable CSV decks from Heisig data sources.

Outputs:
  RTH_deck.csv  — Traditional Hanzi only
  RSH_deck.csv  — Simplified Hanzi only
  RTK_deck.csv  — Kanji only
  Ultimate_deck.csv — All 3 merged, one card per unique character
"""
import csv
import json
import re
import openpyxl
from collections import defaultdict

# ── Paths ──────────────────────────────────────────────────────────────
EXCEL = "data/Heisig's Remembering the Kanji vs. Hanzi v27.xlsx"
RSH_JSON = "data/rsh_parsed.json"
IDS_TXT = "data/IDS.TXT"
UNIFIED_MAP = "data/unified_mapping.json"
HUMAN_REVIEW = "data/unmapped_human_reviewed.csv"

# ── IDS operator labels ────────────────────────────────────────────────
IDS_LABELS = {
    "⿰": "left-right",
    "⿱": "top-bottom",
    "⿲": "left-mid-right",
    "⿳": "top-mid-bottom",
    "⿴": "surround",
    "⿵": "surround-open-bottom",
    "⿶": "surround-open-top",
    "⿷": "surround-open-right",
    "⿸": "top-left-wrap",
    "⿹": "top-right-wrap",
    "⿺": "bottom-left-wrap",
    "⿻": "overlaid",
}
IDS_OPERATORS = set("⿰⿱⿲⿳⿴⿵⿶⿷⿸⿹⿺⿻⿼⿽⿾⿿〾")

# ══════════════════════════════════════════════════════════════════════
# 1. Load data sources
# ══════════════════════════════════════════════════════════════════════

# ── RSH parsed JSON ────────────────────────────────────────────────────
with open(RSH_JSON, "r", encoding="utf-8") as f:
    rsh = json.load(f)

heisig_by_char = {}
heisig_by_keyword = {}
# Two passes: keywords first, then aliases (aliases override keywords)
for e in rsh["characters"] + rsh["primitives"]:
    heisig_by_char[e["character"]] = e
    heisig_by_keyword[e["keyword"]] = e["character"]
for e in rsh["characters"] + rsh["primitives"]:
    for a in e["primitive_aliases"]:
        heisig_by_keyword[a] = e["character"]

# ── Unified mapping (char -> name) ────────────────────────────────────
with open(UNIFIED_MAP, "r", encoding="utf-8") as f:
    unified = json.load(f)

# ── Human review CSV (extra name overrides) ────────────────────────────
human_names = {}
with open(HUMAN_REVIEW, "r", encoding="utf-8") as f:
    reader = csv.DictReader(f)
    for row in reader:
        comp = row["component"]
        name = row.get("your_heisig_name", "").strip()
        if name:
            human_names[comp] = name

# ── Radical mappings (same as recursive_decompose_v2) ──────────────────
radical_map = {
    '訁': '言', '糹': '糸', '釒': '金', '𥫗': '竹', '刂': '刀',
    '彳': '行', '𤣩': '玉', '𧾷': '足', '罒': '网', '乚': '乙',
    '飠': '食', '爫': '爪', '虍': '虎', '𧘇': '衣', '龶': '生',
    '𦍌': '羊', '亍': '行', '牜': '牛', '覀': '西', '丬': '爿',
    '䒑': '丷', '亻': '人', '氵': '水', '扌': '手', '忄': '心',
    '犭': '犬', '礻': '示', '衤': '衣', '灬': '火', '⺌': '小',
    '⺊': '卜', '讠': '言', '钅': '金', '饣': '食', '纟': '糸',
    '贝': '貝', '车': '車', '见': '見', '门': '門', '鱼': '魚',
    '马': '馬', '鸟': '鳥', '页': '頁', '风': '風', '⺝': '月',
    '⺼': '月', '⺶': '羊', '⺀': '八', '⺄': '乙', '⺆': '冂',
    '⺈': '刀',
}
for variant, parent in radical_map.items():
    if variant not in heisig_by_char and parent in heisig_by_char:
        heisig_by_char[variant] = {
            "character": variant,
            "keyword": heisig_by_char[parent]["keyword"],
            "type": "radical_variant",
            "primitive_aliases": heisig_by_char[parent]["primitive_aliases"],
            "components": [],
            "variant_of": parent,
        }

# ── Excel workbook ─────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL)
ws = wb["RTH+RSH+RTK"]

# Trad/Kanji -> Simplified variant mapping into heisig_by_char
for row in ws.iter_rows(min_row=2, values_only=True):
    th, sh, k = row[3], row[4], row[5]
    for char in [th, k]:
        if char and sh and char != sh and char not in heisig_by_char and sh in heisig_by_char:
            heisig_by_char[char] = {
                "character": char,
                "keyword": heisig_by_char[sh]["keyword"],
                "type": "variant",
                "primitive_aliases": heisig_by_char[sh]["primitive_aliases"],
                "components": [],
                "variant_of": sh,
            }

# ── IDS data ───────────────────────────────────────────────────────────
# Numbered component definitions
numbered_components = {}
with open(IDS_TXT, "r", encoding="utf-8-sig") as f:
    for line in f:
        m = re.match(r'^#\s+\{(\d+)\}\s+(.+)', line)
        if m:
            num = int(m.group(1))
            desc = m.group(2).strip()
            parts = desc.split('\t')
            description = parts[0].strip()
            expansion = parts[-1].strip() if len(parts) > 1 else None
            if expansion == '？':
                expansion = None
            numbered_components[num] = {"description": description, "expansion": expansion}

# IDS sequences
ids_map = {}
with open(IDS_TXT, "r", encoding="utf-8-sig") as f:
    for line in f:
        if line.startswith("#") or line.strip() == "":
            continue
        parts = line.strip().split("\t")
        if len(parts) >= 3:
            ids_map[parts[1]] = parts[2:]

# ── CC-CEDICT from Excel ──────────────────────────────────────────────
ws_cedict = wb["CC-CEDICT"]
# Group readings by (TH, SH) pair
cedict_by_char = defaultdict(list)  # char -> [(pinyin, definition)]
for row in ws_cedict.iter_rows(min_row=2, values_only=True):
    th, sh, pinyin, defn = row[0], row[1], row[2], row[3]
    if not pinyin:
        continue
    for char in [th, sh]:
        if char:
            cedict_by_char[char].append((pinyin, defn or ""))

print(f"Loaded: {len(heisig_by_char)} Heisig chars, {len(ids_map)} IDS entries, "
      f"{len(cedict_by_char)} CC-CEDICT chars, {len(unified)} unified mappings")

# ══════════════════════════════════════════════════════════════════════
# 2. Helper functions
# ══════════════════════════════════════════════════════════════════════

def get_heisig_name(char):
    """Get the Heisig primitive/keyword name for a character."""
    if char in human_names:
        return human_names[char]
    if char in heisig_by_char:
        e = heisig_by_char[char]
        if e["primitive_aliases"]:
            return e["primitive_aliases"][0]
        return e["keyword"]
    if char in unified:
        return unified[char]["name"]
    return None


def tokenize_ids(ids_str):
    cleaned = re.sub(r'\$\([^)]*\)', '', ids_str)
    cleaned = cleaned.replace('^', '').strip()
    tokens = []
    i = 0
    while i < len(cleaned):
        ch = cleaned[i]
        if ch == '{':
            j = cleaned.index('}', i)
            tokens.append(('numbered', int(cleaned[i+1:j])))
            i = j + 1
        elif ch in ' \t\r\n()':
            i += 1
        else:
            tokens.append(('char', ch))
            i += 1
    return tokens


def parse_ids(ids_str):
    tokens = tokenize_ids(ids_str)
    pos = [0]
    def parse_next():
        if pos[0] >= len(tokens):
            return None
        tok_type, tok_val = tokens[pos[0]]
        pos[0] += 1
        if tok_type == 'char' and tok_val in IDS_OPERATORS:
            n = 3 if tok_val in '⿲⿳' else (1 if tok_val == '〾' else 2)
            children = []
            for _ in range(n):
                child = parse_next()
                if child is not None:
                    children.append(child)
            return ('op', tok_val, children)
        elif tok_type == 'numbered':
            return ('numbered', tok_val)
        elif tok_type == 'char':
            return ('char', tok_val)
        return parse_next()
    return parse_next()


def recursive_decompose(char, depth=0, max_depth=10, seen=None):
    """Decompose a character into named Heisig components."""
    if seen is None:
        seen = set()
    if char in seen or depth > max_depth:
        return {"char": char, "name": get_heisig_name(char)}
    seen = seen | {char}
    name = get_heisig_name(char)

    # Heisig decomposition from XML (direct only — not via variant,
    # since trad/kanji variants often have different internal structure)
    if char in heisig_by_char and heisig_by_char[char].get("components"):
        e = heisig_by_char[char]
        children = []
        for comp_name in e["components"]:
            comp_char = heisig_by_keyword.get(comp_name)
            children.append({"char": comp_char or "?", "name": comp_name})
        return {"char": char, "name": name, "source": "heisig", "children": children}

    # For variants without their own components, prefer IDS over
    # the simplified decomposition (which may be structurally wrong)
    if char in ids_map:
        tree = parse_ids(ids_map[char][0])
        return _decompose_ids_tree(tree, depth, max_depth, seen)

    # If no IDS available, try the variant's decomposition as fallback
    if char in heisig_by_char:
        variant_of = heisig_by_char[char].get("variant_of")
        if variant_of and variant_of in heisig_by_char and heisig_by_char[variant_of].get("components"):
            e = heisig_by_char[variant_of]
            children = []
            for comp_name in e["components"]:
                comp_char = heisig_by_keyword.get(comp_name)
                children.append({"char": comp_char or "?", "name": comp_name})
            return {"char": char, "name": name, "source": "heisig_variant", "children": children}

    # Heisig atomic or mapped — stop
    if name:
        return {"char": char, "name": name, "source": "heisig_atomic"}

    return {"char": char, "name": None, "source": "unknown"}


def _decompose_ids_tree(tree, depth, max_depth, seen):
    if tree is None:
        return {"char": "?", "name": None, "source": "parse_error"}
    kind = tree[0]
    if kind == 'char':
        return recursive_decompose(tree[1], depth + 1, max_depth, seen)
    elif kind == 'numbered':
        num = tree[1]
        comp = numbered_components.get(num, {})
        expansion = comp.get("expansion")
        if expansion:
            subtree = parse_ids(expansion)
            return _decompose_ids_tree(subtree, depth + 1, max_depth, seen)
        return {"char": f"{{{num}}}", "name": None, "source": "numbered_component"}
    elif kind == 'op':
        op = tree[1]
        children = [_decompose_ids_tree(c, depth + 1, max_depth, seen) for c in tree[2]]
        return {"char": "".join(c.get("char", "?") for c in children),
                "name": None, "source": "ids", "operator": op, "children": children}
    return {"char": "?", "name": None, "source": "parse_error"}


def collect_leaves(node, is_root=True):
    """Get named component names from a decomposition tree.
    Stops at any named node (doesn't recurse into its sub-components)."""
    # Named non-root node: stop here
    if not is_root and node.get("name"):
        return [node["name"]]
    if "children" not in node:
        name = node.get("name")
        return [name] if name else [node.get("char", "?")]
    result = []
    for child in node["children"]:
        result.extend(collect_leaves(child, is_root=False))
    return result


def collect_leaf_details(node, is_root=True):
    """Get (char, name) pairs for named components.
    Stops at any named node (doesn't recurse into its sub-components)."""
    if not is_root and node.get("name"):
        return [(node.get("char", "?"), node["name"])]
    if "children" not in node:
        return [(node.get("char", "?"), node.get("name"))]
    result = []
    for child in node["children"]:
        result.extend(collect_leaf_details(child, is_root=False))
    return result


def get_raw_ids(char):
    """Get the cleaned raw IDS string for a character."""
    if char in ids_map:
        raw = ids_map[char][0]
        return re.sub(r'\$\([^)]*\)', '', raw).replace('^', '').strip()
    return ""


def get_top_operator(char):
    """Get the top-level IDS operator for a character's spatial layout."""
    cleaned = get_raw_ids(char)
    if cleaned and cleaned[0] in IDS_LABELS:
        op = cleaned[0]
        return f"{op} ({IDS_LABELS[op]})"
    return ""


def format_reading(char):
    """Format CC-CEDICT readings for a character.

    Groups definitions by pinyin reading, limits to 3-4 meanings per reading,
    filters obscure readings.
    """
    entries = cedict_by_char.get(char, [])
    if not entries:
        return ""

    # Group by pinyin base (ignoring tone differences that are the same syllable)
    by_pinyin = defaultdict(list)
    for pinyin, defn in entries:
        by_pinyin[pinyin].append(defn)

    # Filter: skip surname-only or very obscure entries
    # Keep max 4 readings, each with max 4 meanings
    parts = []
    for pinyin, defns in by_pinyin.items():
        # Merge all definitions for this reading
        all_meanings = []
        for d in defns:
            # Split on "/" which CC-CEDICT uses as separator
            for m in d.split("/"):
                m = m.strip()
                if m and m not in all_meanings:
                    all_meanings.append(m)

        # Filter out surname-only entries if there are other readings
        if len(by_pinyin) > 1 and len(all_meanings) <= 2:
            if all(("surname" in m.lower() or "name" in m.lower()) for m in all_meanings):
                continue

        # Limit meanings
        shown = all_meanings[:4]
        parts.append(f"{pinyin}: {', '.join(shown)}")

        if len(parts) >= 4:
            break

    return " | ".join(parts)


# ══════════════════════════════════════════════════════════════════════
# 3. Build card data from Excel
# ══════════════════════════════════════════════════════════════════════

# Collect all Excel row data
excel_rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    excel_rows.append(row)

# Build card dicts keyed by character
cards = {}  # char -> card dict

def ensure_card(char):
    if char not in cards:
        cards[char] = {
            "character": char,
            "keyword": "",
            "RTH_number": "",
            "RSH_number": "",
            "RTK_number": "",
            "reading": "",
            "decomposition": "",
            "spatial": "",
            "components_detail": "",
            "tags": [],
            "books": set(),  # internal tracking
        }
    return cards[char]


for row in excel_rows:
    rth_num, rsh_num, rtk_num = row[0], row[1], row[2]
    th, sh, k = row[3], row[4], row[5]
    rth_kw, rsh_kw, rtk_kw = row[7], row[8], row[9]
    rth_read, th_read = row[10], row[11]
    rth_lesson, rsh_lesson, rtk_lesson = row[12], row[13], row[14]

    # Parse RSH number (can be "ch # 0041" format or numeric)
    def parse_num(val):
        if val is None:
            return None
        if isinstance(val, (int, float)):
            return int(val)
        m = re.match(r'ch\s*#\s*(\d+)', str(val))
        if m:
            return int(m.group(1))
        return None

    rth_n = parse_num(rth_num)
    rsh_n = parse_num(rsh_num)
    rtk_n = parse_num(rtk_num)

    # Process each book's character
    book_entries = [
        (th, rth_n, rth_kw, rth_lesson, "RTH"),
        (sh, rsh_n, rsh_kw, rsh_lesson, "RSH"),
        (k, rtk_n, rtk_kw, rtk_lesson, "RTK"),
    ]

    for char, num, kw, lesson, book in book_entries:
        if not char:
            continue
        card = ensure_card(char)

        if num:
            card["books"].add(book)
            card[f"{book}_number"] = str(num)

        if kw and not card["keyword"]:
            card["keyword"] = kw
        elif kw and card["keyword"] and kw != card["keyword"]:
            # Append alternate keyword if different
            if kw not in card["keyword"]:
                card["keyword"] += f" / {kw}"

        if lesson:
            # Convert "RSH1-L01" -> "RSH1::L01" for Anki nested tags
            card["tags"].append(lesson.replace("-", "::"))


# ══════════════════════════════════════════════════════════════════════
# 4. Enrich cards: decomposition, spatial, readings, primitives
# ══════════════════════════════════════════════════════════════════════

for char, card in cards.items():
    # Deck column: which book(s) this character belongs to
    card["deck"] = " ".join(sorted(card["books"])) if card["books"] else ""

    # Reading from CC-CEDICT (skip for RTK-only / kanji-only)
    if "RTH" in card["books"] or "RSH" in card["books"]:
        card["reading"] = format_reading(char)

    # Decomposition + components_detail
    tree = recursive_decompose(char)
    leaves = collect_leaves(tree)
    leaf_details = collect_leaf_details(tree)

    if tree.get("source") == "heisig" or (tree.get("children") and len(leaves) > 0):
        card["decomposition"] = " + ".join(leaves)
        # components_detail: each component char = keyword (alias: X if used differently)
        detail_parts = []
        seen_detail = set()
        for ch, decomp_name in leaf_details:
            if decomp_name and ch != "?" and ch not in seen_detail:
                seen_detail.add(ch)
                # Get the character's actual keyword
                if ch in heisig_by_char:
                    keyword = heisig_by_char[ch].get("keyword", decomp_name)
                    aliases = heisig_by_char[ch].get("primitive_aliases", [])
                else:
                    keyword = decomp_name
                    aliases = []
                part = f"{ch} = {keyword}"
                # If decomposition used a different name (an alias), note it
                if decomp_name.lower() != keyword.lower() and decomp_name.lower() in [a.lower() for a in aliases]:
                    part += f" (alias: {decomp_name})"
                elif aliases and any(a.lower() != keyword.lower() for a in aliases):
                    # Show other aliases even if not used in this decomposition
                    other_aliases = [a for a in aliases if a.lower() != keyword.lower()]
                    if other_aliases:
                        part += f" (alias: {', '.join(other_aliases)})"
                detail_parts.append(part)
        card["components_detail"] = ", ".join(detail_parts)
    elif tree.get("source") == "heisig_atomic":
        card["decomposition"] = ""  # atomic, no sub-components
        card["components_detail"] = ""

    # Spatial from IDS
    card["spatial"] = get_top_operator(char)

    # Raw IDS string
    card["ids"] = get_raw_ids(char)

    # Tags: add primitive tag if applicable
    if char in heisig_by_char:
        entry = heisig_by_char[char]
        if entry.get("type") == "primitive" or entry.get("primitive_aliases"):
            card["tags"].append("primitive")

    card["tags"] = " ".join(sorted(set(card["tags"])))


# ══════════════════════════════════════════════════════════════════════
# 5. Add standalone primitive cards
# ══════════════════════════════════════════════════════════════════════

primitives_added = 0
for entry in rsh["primitives"]:
    char = entry["character"]
    if char in cards:
        continue  # already covered
    # Standalone primitive not in any book's character list
    tree = recursive_decompose(char)
    leaves = collect_leaves(tree)
    leaf_details = collect_leaf_details(tree)

    aliases = entry["primitive_aliases"]
    kw = entry["keyword"]
    alias_str = f" (also: {', '.join(aliases)})" if aliases else ""

    card = {
        "character": char,
        "keyword": f"{kw}{alias_str}",
        "RTH_number": "",
        "RSH_number": "",
        "RTK_number": "",
        "reading": format_reading(char),
        "decomposition": " + ".join(leaves) if tree.get("children") else "",
        "spatial": get_top_operator(char),
        "ids": get_raw_ids(char),
        "components_detail": "",
        "tags": "primitive",
    }
    if tree.get("children"):
        detail_parts = []
        seen_detail = set()
        for ch, decomp_name in leaf_details:
            if decomp_name and ch != "?" and ch not in seen_detail:
                seen_detail.add(ch)
                if ch in heisig_by_char:
                    keyword = heisig_by_char[ch].get("keyword", decomp_name)
                    aliases = heisig_by_char[ch].get("primitive_aliases", [])
                else:
                    keyword = decomp_name
                    aliases = []
                part = f"{ch} = {keyword}"
                if decomp_name.lower() != keyword.lower() and decomp_name.lower() in [a.lower() for a in aliases]:
                    part += f" (alias: {decomp_name})"
                elif aliases and any(a.lower() != keyword.lower() for a in aliases):
                    other_aliases = [a for a in aliases if a.lower() != keyword.lower()]
                    if other_aliases:
                        part += f" (alias: {', '.join(other_aliases)})"
                detail_parts.append(part)
        card["components_detail"] = ", ".join(detail_parts)

    cards[char] = card
    primitives_added += 1

# Also add character-primitives that might only be in RSH XML but not Excel
for entry in rsh["characters"]:
    char = entry["character"]
    if char in cards:
        continue
    if not entry["primitive_aliases"]:
        continue
    tree = recursive_decompose(char)
    leaves = collect_leaves(tree)
    leaf_details = collect_leaf_details(tree)

    card = {
        "character": char,
        "keyword": entry["keyword"],
        "RTH_number": "",
        "RSH_number": str(entry["number"]) if entry.get("number") else "",
        "RTK_number": "",
        "reading": format_reading(char),
        "decomposition": " + ".join(leaves) if tree.get("children") else "",
        "spatial": get_top_operator(char),
        "ids": get_raw_ids(char),
        "components_detail": "",
        "tags": "primitive",
    }
    if tree.get("children"):
        detail_parts = []
        seen_detail = set()
        for ch, decomp_name in leaf_details:
            if decomp_name and ch != "?" and ch not in seen_detail:
                seen_detail.add(ch)
                if ch in heisig_by_char:
                    keyword = heisig_by_char[ch].get("keyword", decomp_name)
                    aliases = heisig_by_char[ch].get("primitive_aliases", [])
                else:
                    keyword = decomp_name
                    aliases = []
                part = f"{ch} = {keyword}"
                if decomp_name.lower() != keyword.lower() and decomp_name.lower() in [a.lower() for a in aliases]:
                    part += f" (alias: {decomp_name})"
                elif aliases and any(a.lower() != keyword.lower() for a in aliases):
                    other_aliases = [a for a in aliases if a.lower() != keyword.lower()]
                    if other_aliases:
                        part += f" (alias: {', '.join(other_aliases)})"
                detail_parts.append(part)
        card["components_detail"] = ", ".join(detail_parts)

    cards[char] = card
    primitives_added += 1

print(f"Standalone primitives added: {primitives_added}")

# ══════════════════════════════════════════════════════════════════════
# 6. Output CSVs
# ══════════════════════════════════════════════════════════════════════

COLUMNS = ["character", "keyword", "RTH_number", "RSH_number", "RTK_number",
           "reading", "decomposition", "spatial", "ids", "components_detail", "deck", "tags"]


def write_deck(filename, filter_fn):
    """Write a CSV deck, filtering cards by filter_fn."""
    rows = []
    for char, card in sorted(cards.items(), key=lambda x: x[0]):
        if filter_fn(card):
            rows.append([card.get(col, "") for col in COLUMNS])

    with open(filename, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(COLUMNS)
        writer.writerows(rows)

    return len(rows)


# Collect which characters belong to which books from Excel
rth_chars = set()
rsh_chars = set()
rtk_chars = set()
for row in excel_rows:
    if row[0] and row[3]:  # RTH number + TH char
        rth_chars.add(row[3])
    if row[1] and row[4]:  # RSH number + SH char
        rsh_chars.add(row[4])
    if row[2] and row[5]:  # RTK number + K char
        rtk_chars.add(row[5])

is_primitive = lambda c: "primitive" in c.get("tags", "")
n_rth = write_deck("RTH_deck.csv", lambda c: c["character"] in rth_chars or is_primitive(c))
n_rsh = write_deck("RSH_deck.csv", lambda c: c["character"] in rsh_chars or is_primitive(c))
n_rtk = write_deck("RTK_deck.csv", lambda c: c["character"] in rtk_chars or is_primitive(c))
n_ult = write_deck("Ultimate_deck.csv", lambda c: True)

# ══════════════════════════════════════════════════════════════════════
# 7. Summary
# ══════════════════════════════════════════════════════════════════════

print(f"\n{'='*60}")
print(f"  RTH_deck.csv:      {n_rth} cards")
print(f"  RSH_deck.csv:      {n_rsh} cards")
print(f"  RTK_deck.csv:      {n_rtk} cards")
print(f"  Ultimate_deck.csv: {n_ult} cards")
print(f"{'='*60}")

# Spot checks
for ch, expected in [("虎", "magic wand"), ("國", "pent in")]:
    card = cards.get(ch)
    if card:
        decomp = card["decomposition"]
        has = expected in decomp if decomp else False
        status = "OK" if has else "MISSING"
        print(f"  Spot check {ch}: decomposition = '{decomp}' [{status}]")

# Primitive count
prim_count = sum(1 for c in cards.values() if "primitive" in c.get("tags", ""))
print(f"  Primitive-tagged cards: {prim_count}")
print(f"  Total unique characters: {len(cards)}")
