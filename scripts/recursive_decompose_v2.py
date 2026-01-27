"""Recursive decomposition v2: fix IDS parser to handle {N} components."""
import json
import re
import openpyxl
from collections import Counter

# --- Load IDS numbered component definitions ---
numbered_components = {}  # {N} -> description string
with open("data/IDS.TXT", "r", encoding="utf-8-sig") as f:
    for line in f:
        m = re.match(r'^#\s+\{(\d+)\}\s+(.+)', line)
        if m:
            num = int(m.group(1))
            desc = m.group(2).strip()
            # Extract the IDS expansion if given (last column after tabs)
            parts = desc.split('\t')
            description = parts[0].strip()
            expansion = parts[-1].strip() if len(parts) > 1 else None
            if expansion == '？':
                expansion = None
            numbered_components[num] = {
                "description": description,
                "expansion": expansion,
            }

print(f"Numbered IDS components loaded: {len(numbered_components)}")
# Show a few
for n in [73, 87, 88, 13, 28]:
    c = numbered_components.get(n, {})
    print(f"  {{{n}}}: {c.get('description', '?')} -> {c.get('expansion', 'no expansion')}")

# --- Load RSH data ---
with open("data/rsh_parsed.json", "r", encoding="utf-8") as f:
    rsh = json.load(f)

heisig_by_char = {}
heisig_by_keyword = {}
for e in rsh["characters"] + rsh["primitives"]:
    heisig_by_char[e["character"]] = e
    heisig_by_keyword[e["keyword"]] = e["character"]
    for a in e["primitive_aliases"]:
        heisig_by_keyword[a] = e["character"]

# Radical mappings
radical_map = {
    '訁': '言', '糹': '糸', '釒': '金', '𥫗': '竹', '刂': '刀',
    '彳': '行', '𤣩': '玉', '𧾷': '足', '罒': '网', '乚': '乙',
    '飠': '食', '爫': '爪', '虍': '虎', '𧘇': '衣', '龶': '生',
    '𦍌': '羊', '亍': '行', '牜': '牛', '覀': '西', '丬': '爿',
    '䒑': '丷', '亻': '人', '氵': '水', '扌': '手', '忄': '心',
    '犭': '犬', '礻': '示', '衤': '衣', '灬': '火', '⺌': '小',
    '⺊': '卜', '讠': '言', '钅': '金', '饣': '食', '纟': '糸',
    '贝': '貝', '车': '車', '见': '見', '门': '門', '鱼': '魚',
    '马': '馬', '鸟': '鳥', '页': '頁', '风': '風', '⺝': '月',
    '⺼': '月', '⺶': '羊', '⺀': '八', '⺄': '乙', '⺆': '冂',
    '⺈': '刀',
}
for variant, parent in radical_map.items():
    if variant not in heisig_by_char and parent in heisig_by_char:
        heisig_by_char[variant] = {
            "character": variant,
            "keyword": heisig_by_char[parent]["keyword"],
            "type": "radical_variant",
            "primitive_aliases": heisig_by_char[parent]["primitive_aliases"],
            "components": [],
            "variant_of": parent,
        }

# Trad/Kanji -> Simplified mappings
wb = openpyxl.load_workbook("data/Heisig's Remembering the Kanji vs. Hanzi v27.xlsx")
ws = wb["RTH+RSH+RTK"]
for row in ws.iter_rows(min_row=2, values_only=True):
    th, sh, k = row[3], row[4], row[5]
    for char in [th, k]:
        if char and sh and char != sh and char not in heisig_by_char and sh in heisig_by_char:
            heisig_by_char[char] = {
                "character": char,
                "keyword": heisig_by_char[sh]["keyword"],
                "type": "variant",
                "primitive_aliases": heisig_by_char[sh]["primitive_aliases"],
                "components": [],
                "variant_of": sh,
            }

print(f"Total Heisig lookup entries: {len(heisig_by_char)}")

# --- Load IDS ---
ids_map = {}
with open("data/IDS.TXT", "r", encoding="utf-8-sig") as f:
    for line in f:
        if line.startswith("#") or line.strip() == "":
            continue
        parts = line.strip().split("\t")
        if len(parts) >= 3:
            ids_map[parts[1]] = parts[2:]

IDS_OPERATORS = set("⿰⿱⿲⿳⿴⿵⿶⿷⿸⿹⿺⿻⿼⿽⿾⿿〾")

def tokenize_ids(ids_str):
    """Tokenize IDS string, handling {N} numbered components."""
    cleaned = re.sub(r'\$\([^)]*\)', '', ids_str)
    cleaned = cleaned.replace('^', '').strip()
    tokens = []
    i = 0
    while i < len(cleaned):
        ch = cleaned[i]
        if ch == '{':
            # Find closing brace
            j = cleaned.index('}', i)
            num = int(cleaned[i+1:j])
            tokens.append(('numbered', num))
            i = j + 1
        elif ch in ' \t\r\n()':
            i += 1
        else:
            tokens.append(('char', ch))
            i += 1
    return tokens


def parse_ids(ids_str):
    """Parse IDS string into tree, handling {N} components."""
    tokens = tokenize_ids(ids_str)
    pos = [0]

    def parse_next():
        if pos[0] >= len(tokens):
            return None
        tok_type, tok_val = tokens[pos[0]]
        pos[0] += 1

        if tok_type == 'char' and tok_val in IDS_OPERATORS:
            n = 3 if tok_val in '⿲⿳' else (1 if tok_val == '〾' else 2)
            children = []
            for _ in range(n):
                child = parse_next()
                if child is not None:
                    children.append(child)
            return ('op', tok_val, children)
        elif tok_type == 'numbered':
            return ('numbered', tok_val)
        elif tok_type == 'char':
            return ('char', tok_val)
        return parse_next()

    return parse_next()


def get_heisig_name(char):
    if char in heisig_by_char:
        e = heisig_by_char[char]
        if e["primitive_aliases"]:
            return e["primitive_aliases"][0]
        return e["keyword"]
    return None


def recursive_decompose(char, depth=0, max_depth=10, seen=None):
    if seen is None:
        seen = set()
    if char in seen or depth > max_depth:
        return {"char": char, "name": get_heisig_name(char)}
    seen = seen | {char}

    name = get_heisig_name(char)

    # Strategy 1: Heisig decomposition from XML
    if char in heisig_by_char and heisig_by_char[char].get("components"):
        e = heisig_by_char[char]
        children = []
        for comp_name in e["components"]:
            comp_char = heisig_by_keyword.get(comp_name)
            if comp_char:
                children.append({"char": comp_char, "name": comp_name})
            else:
                children.append({"char": "?", "name": comp_name})
        return {"char": char, "name": name, "source": "heisig", "children": children}

    # Strategy 2: Heisig atomic (named but no decomposition)
    if name and char in heisig_by_char and not heisig_by_char[char].get("components"):
        return {"char": char, "name": name, "source": "heisig_atomic"}

    # Strategy 3: Named via radical/variant map, stop here
    if name:
        return {"char": char, "name": name, "source": "mapped"}

    # Strategy 4: IDS decomposition, recurse
    if char in ids_map:
        tree = parse_ids(ids_map[char][0])
        return _decompose_ids_tree(tree, depth, max_depth, seen)

    return {"char": char, "name": None, "source": "unknown"}


def _decompose_ids_tree(tree, depth, max_depth, seen):
    if tree is None:
        return {"char": "?", "name": None, "source": "parse_error"}

    if isinstance(tree, tuple):
        kind = tree[0]
        if kind == 'char':
            ch = tree[1]
            return recursive_decompose(ch, depth + 1, max_depth, seen)
        elif kind == 'numbered':
            num = tree[1]
            comp = numbered_components.get(num, {})
            desc = comp.get("description", f"component {num}")
            expansion = comp.get("expansion")
            # Try to expand the numbered component via its IDS expansion
            if expansion:
                subtree = parse_ids(expansion)
                result = _decompose_ids_tree(subtree, depth + 1, max_depth, seen)
                if result.get("name") is None:
                    result["note"] = f"{{{num}}}: {desc}"
                return result
            return {
                "char": f"{{{num}}}",
                "name": None,
                "source": "numbered_component",
                "note": desc,
            }
        elif kind == 'op':
            op = tree[1]
            children = tree[2]
            decomposed = []
            for child in children:
                decomposed.append(_decompose_ids_tree(child, depth + 1, max_depth, seen))
            # If all children are resolved, we're good
            char_str = "".join(
                c.get("char", "?") for c in decomposed
            )
            return {
                "char": char_str,
                "name": None,
                "source": "ids",
                "operator": op,
                "children": decomposed,
            }

    return {"char": "?", "name": None, "source": "parse_error"}


def format_tree(node, indent=0):
    prefix = "  " * indent
    char = node.get("char", "?")
    name = node.get("name") or "???"
    source = node.get("source", "")
    note = node.get("note", "")
    variant = ""
    if char in heisig_by_char and "variant_of" in heisig_by_char[char]:
        variant = f" (variant of {heisig_by_char[char]['variant_of']})"

    label = f"{char} [{name}]{variant}"
    if note:
        label += f" -- {note}"

    if "children" in node:
        op = node.get("operator", "→")
        print(f"{prefix}{label} ({source}) {op}")
        for child in node["children"]:
            format_tree(child, indent + 1)
    else:
        print(f"{prefix}{label} ({source})")


def count_leaves(node):
    if "children" not in node:
        return (1, 0) if node.get("name") else (0, 1)
    resolved = unresolved = 0
    for child in node["children"]:
        r, u = count_leaves(child)
        resolved += r
        unresolved += u
    return resolved, unresolved


# Load all Excel characters
all_excel = set()
excel_info = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    for col, book in [(3, 'RTH'), (4, 'RSH'), (5, 'RTK')]:
        if row[col]:
            all_excel.add(row[col])
            if row[col] not in excel_info:
                excel_info[row[col]] = []
            kw_col = {3: 7, 4: 8, 5: 9}[col]
            if row[kw_col]:
                excel_info[row[col]].append(f"{book}: {row[kw_col]}")

# Test examples
test_chars = ['虎', '慮', '龍', '鬱', '飛', '愛', '學', '國', '聽', '體',
              '直', '具', '黑', '合', '亭', '帝', '武', '書']
print("\n=== Sample Decompositions ===\n")
for ch in test_chars:
    tree = recursive_decompose(ch)
    format_tree(tree)
    r, u = count_leaves(tree)
    status = "COMPLETE" if u == 0 else f"{u} unresolved"
    print(f"  -> {status}\n")

# Full stats
fully_resolved = 0
partial = 0
unresolved_list = []
leaf_freq = Counter()

for char in sorted(all_excel):
    tree = recursive_decompose(char)
    r, u = count_leaves(tree)
    if u == 0:
        fully_resolved += 1
    else:
        if r > 0:
            partial += 1
        else:
            unresolved_list.append(char)

        def collect(node):
            if "children" not in node:
                if not node.get("name"):
                    key = node.get("char", "?")
                    note = node.get("note", "")
                    leaf_freq[(key, note)] += 1
            else:
                for child in node["children"]:
                    collect(child)
        collect(tree)

print(f"\n=== Coverage: {len(all_excel)} characters ===")
print(f"Fully resolved:     {fully_resolved}")
print(f"Partially resolved: {partial}")
print(f"Unresolved:         {len(unresolved_list)}")

print(f"\nUnresolved leaf components ({len(leaf_freq)}):")
print(f"{'Char':<8} {'Count':>5}  Description")
print("-" * 60)
for (char, note), count in leaf_freq.most_common(40):
    desc = note if note else char
    # Try to find example characters that use this component
    print(f"{char:<8} {count:>5}x  {desc}")

# Show example characters for top unresolved
print("\n=== Examples for top unresolved components ===")
for (leaf_char, note), count in leaf_freq.most_common(15):
    examples = []
    for char in sorted(all_excel):
        tree = recursive_decompose(char)
        def has_leaf(node, target_char, target_note):
            if "children" not in node:
                return not node.get("name") and node.get("char") == target_char
            return any(has_leaf(c, target_char, target_note) for c in node["children"])
        if has_leaf(tree, leaf_char, note):
            info = ", ".join(excel_info.get(char, []))
            examples.append(f"{char}({info})")
            if len(examples) >= 5:
                break
    desc = note if note else leaf_char
    print(f"\n  {leaf_char} [{desc}] - {count}x")
    print(f"    e.g.: {', '.join(examples)}")
