#!/usr/bin/env python3
"""Generate unique keywords for characters using CC-CEDICT and Unihan.

Strategy:
1. Try each CC-CEDICT definition (separated by /) in order
2. Skip definitions that are just variants/references
3. Fall back to Unihan if no CC-CEDICT entry
4. Only use numbered suffixes as last resort
"""

import json
import re
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
CEDICT_PATH = ROOT / "data" / "cedict.txt"
UNIHAN_PATH = ROOT / "data" / "Unihan_Readings.txt"


def parse_cedict():
    """Parse CC-CEDICT into char -> list of definitions."""
    cedict = {}
    if not CEDICT_PATH.exists():
        return cedict

    with open(CEDICT_PATH, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            # Format: 傳統 简体 [pinyin] /def1/def2/
            match = re.match(r"(\S+)\s+(\S+)\s+\[([^\]]+)\]\s+/(.+)/", line)
            if not match:
                continue
            trad, simp, pinyin, defs_str = match.groups()

            # Split definitions and clean them
            defs = []
            for d in defs_str.split("/"):
                d = d.strip()
                if not d:
                    continue
                # Skip pure variant references
                if re.match(r"^(variant of|see |also written|abbr\. for|same as)", d, re.I):
                    continue
                # Clean up parenthetical notes at start but keep the main def
                d = re.sub(r"^\([^)]+\)\s*", "", d)
                if not d:
                    continue
                # Also split on semicolons (e.g., "to gnaw; to bite")
                for sub in d.split(";"):
                    sub = sub.strip()
                    if sub:
                        defs.append(sub)

            # Store for both trad and simp
            for char in [trad, simp]:
                if len(char) == 1:  # Single character only
                    if char not in cedict:
                        cedict[char] = {"pinyin": pinyin, "defs": []}
                    # Add new definitions we haven't seen
                    for d in defs:
                        if d not in cedict[char]["defs"]:
                            cedict[char]["defs"].append(d)

    return cedict


def parse_unihan():
    """Parse Unihan_Readings.txt for kDefinition."""
    unihan = {}
    if not UNIHAN_PATH.exists():
        return unihan

    with open(UNIHAN_PATH, "r", encoding="utf-8") as f:
        for line in f:
            if "\tkDefinition\t" in line:
                parts = line.strip().split("\t")
                if len(parts) >= 3:
                    codepoint = parts[0]  # U+XXXX
                    definition = parts[2]
                    # Convert codepoint to character
                    char = chr(int(codepoint[2:], 16))
                    # Split definitions by semicolon or comma
                    defs = [d.strip() for d in re.split(r"[;,]", definition) if d.strip()]
                    unihan[char] = defs

    return unihan


def normalize_keyword(kw):
    """Normalize keyword for uniqueness checking."""
    kw = kw.lower().strip()
    # Handle singular/plural
    if kw.endswith("ies"):
        kw = kw[:-3] + "y"
    elif kw.endswith("es"):
        kw = kw[:-2]
    elif kw.endswith("s") and not kw.endswith("ss"):
        kw = kw[:-1]
    return kw


def generate_keywords(chars_to_process, existing_keywords):
    """
    Generate unique keywords for a list of characters.

    Args:
        chars_to_process: list of dicts with 'char', 'ids', 'components_detail', 'tags'
        existing_keywords: set of already-used keywords (normalized)

    Returns:
        list of dicts with 'keyword', 'reading', 'source' added
    """
    cedict = parse_cedict()
    unihan = parse_unihan()

    used_keywords = set(existing_keywords)
    results = []
    failed = []

    for entry in chars_to_process:
        char = entry["char"]
        keyword = None
        reading = ""
        source = None

        # Try CC-CEDICT definitions
        if char in cedict:
            reading = cedict[char]["pinyin"]
            for d in cedict[char]["defs"]:
                norm = normalize_keyword(d)
                if norm not in used_keywords:
                    keyword = d
                    source = "cedict"
                    used_keywords.add(norm)
                    break

        # Try Unihan definitions
        if keyword is None and char in unihan:
            for d in unihan[char]:
                norm = normalize_keyword(d)
                if norm not in used_keywords:
                    keyword = d
                    source = "unihan"
                    used_keywords.add(norm)
                    break

        # Last resort: numbered suffix using first definition
        if keyword is None:
            base_def = None
            if char in cedict and cedict[char]["defs"]:
                base_def = cedict[char]["defs"][0]
                reading = cedict[char]["pinyin"]
            elif char in unihan and unihan[char]:
                base_def = unihan[char][0]

            if base_def:
                # Find next available number
                for i in range(2, 100):
                    candidate = f"{base_def} ({i})"
                    norm = normalize_keyword(candidate)
                    if norm not in used_keywords:
                        keyword = candidate
                        source = "cedict+num" if char in cedict else "unihan+num"
                        used_keywords.add(norm)
                        break

        if keyword:
            entry["keyword"] = keyword
            entry["reading"] = reading
            entry["source"] = source
            results.append(entry)
        else:
            failed.append(char)

    return results, failed


def load_existing_keywords():
    """Load keywords already in use from heisig_data.json."""
    data_path = ROOT / "heisig_addon" / "data" / "heisig_data.json"
    keywords = set()

    if data_path.exists():
        with open(data_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            for char, info in data.items():
                kw = info.get("keyword", "")
                if kw:
                    keywords.add(normalize_keyword(kw))

    return keywords


def load_ids_data():
    """Load IDS decomposition data."""
    ids_path = ROOT / "data" / "IDS.TXT"
    ids_data = {}
    if ids_path.exists():
        with open(ids_path, "r", encoding="utf-8") as f:
            for line in f:
                if line.startswith("#") or "\t" not in line:
                    continue
                parts = line.strip().split("\t")
                if len(parts) >= 2:
                    char = parts[1]
                    ids = parts[2] if len(parts) > 2 else ""
                    ids_data[char] = ids
    return ids_data


def get_component_detail(char, ids, heisig_data):
    """Build components_detail string for a character."""
    if not ids:
        return ""

    components = []
    seen = set()
    for c in ids:
        if c in seen:
            continue
        if c in heisig_data and c != char:
            kw = heisig_data[c].get("keyword", "")
            if kw:
                components.append(f"{c} = {kw}")
                seen.add(c)

    return ", ".join(components)


def regenerate_all():
    """Regenerate all mainland and taiwan character data with improved keywords."""
    import csv
    import openpyxl

    # Load existing Heisig data
    heisig_path = ROOT / "heisig_addon" / "data" / "heisig_data.json"
    with open(heisig_path, "r", encoding="utf-8") as f:
        heisig_data = json.load(f)

    # Collect existing keywords (from Heisig books only, not auto-generated)
    existing_keywords = set()
    heisig_chars = set()
    for char, info in heisig_data.items():
        # Check if it's an original Heisig character (has book number)
        if info.get("RSH_number") or info.get("RTH_number") or info.get("RTK_number"):
            kw = info.get("keyword", "")
            if kw:
                existing_keywords.add(normalize_keyword(kw))
                heisig_chars.add(char)
        # Also count primitives
        if "primitive" in info.get("tags", ""):
            kw = info.get("keyword", "")
            if kw:
                existing_keywords.add(normalize_keyword(kw))
                heisig_chars.add(char)

    print(f"Loaded {len(heisig_chars)} Heisig characters with {len(existing_keywords)} unique keywords")

    # Load IDS data
    ids_data = load_ids_data()

    # Parse CC-CEDICT and Unihan
    cedict = parse_cedict()
    unihan = parse_unihan()
    print(f"Loaded CC-CEDICT: {len(cedict)} entries, Unihan: {len(unihan)} entries")

    # Load mainland characters by level
    mainland_path = ROOT / "data" / "additional_characters" / "mainland_characters.csv"
    mainland_chars = []
    with open(mainland_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        next(reader)  # skip header
        level = "ML::L1"
        count = 0
        for row in reader:
            if not row:
                continue
            char = row[0].strip()
            if not char or char in heisig_chars:
                continue
            # Determine level based on count (approximate)
            if count < 3500:
                level = "ML::L1"
            elif count < 6500:
                level = "ML::L2"
            else:
                level = "ML::L3"
            count += 1
            mainland_chars.append({
                "char": char,
                "ids": ids_data.get(char, ""),
                "tags": level
            })

    # Load Taiwan characters from A and B sheets only (skip rare C sheet)
    taiwan_path = ROOT / "data" / "additional_characters" / "taiwan_char list.xlsx"
    taiwan_chars = []
    if taiwan_path.exists():
        wb = openpyxl.load_workbook(taiwan_path, read_only=True)
        # Read A sheet (common 4808)
        if "A常用字4808" in wb.sheetnames:
            ws = wb["A常用字4808"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                char = row[1] if len(row) > 1 else row[0]
                if char and isinstance(char, str):
                    char = char.strip()
                    if char and char not in heisig_chars:
                        taiwan_chars.append({
                            "char": char,
                            "ids": ids_data.get(char, ""),
                            "tags": "TW::A"
                        })
        # Read B sheet (secondary 6329)
        if "B次常用字6329" in wb.sheetnames:
            ws = wb["B次常用字6329"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                char = row[1] if len(row) > 1 else row[0]
                if char and isinstance(char, str):
                    char = char.strip()
                    if char and char not in heisig_chars:
                        taiwan_chars.append({
                            "char": char,
                            "ids": ids_data.get(char, ""),
                            "tags": "TW::B"
                        })
        wb.close()

    print(f"Mainland chars to process: {len(mainland_chars)}")
    print(f"Taiwan chars to process: {len(taiwan_chars)}")

    # Generate keywords for mainland
    used_keywords = set(existing_keywords)
    mainland_results = []
    mainland_failed = []

    for entry in mainland_chars:
        char = entry["char"]
        keyword = None
        reading = ""
        source = None

        # Try CC-CEDICT definitions
        if char in cedict:
            reading = cedict[char]["pinyin"]
            for d in cedict[char]["defs"]:
                norm = normalize_keyword(d)
                if norm not in used_keywords:
                    keyword = d
                    source = "cedict"
                    used_keywords.add(norm)
                    break

        # Try Unihan
        if keyword is None and char in unihan:
            for d in unihan[char]:
                norm = normalize_keyword(d)
                if norm not in used_keywords:
                    keyword = d
                    source = "unihan"
                    used_keywords.add(norm)
                    break

        # Last resort: numbered suffix
        if keyword is None:
            base_def = None
            if char in cedict and cedict[char]["defs"]:
                base_def = cedict[char]["defs"][0]
                reading = cedict[char]["pinyin"]
            elif char in unihan and unihan[char]:
                base_def = unihan[char][0]

            if base_def:
                for i in range(2, 100):
                    candidate = f"{base_def} ({i})"
                    norm = normalize_keyword(candidate)
                    if norm not in used_keywords:
                        keyword = candidate
                        source = "numbered"
                        used_keywords.add(norm)
                        break

        if keyword:
            entry["keyword"] = keyword
            entry["reading"] = reading
            entry["source"] = source
            entry["components_detail"] = get_component_detail(char, entry["ids"], heisig_data)
            mainland_results.append(entry)
        else:
            mainland_failed.append(char)

    # Generate keywords for Taiwan (that aren't already in mainland)
    mainland_char_set = {e["char"] for e in mainland_results}
    taiwan_results = []
    taiwan_failed = []

    for entry in taiwan_chars:
        char = entry["char"]
        if char in mainland_char_set:
            continue

        keyword = None
        reading = ""
        source = None

        if char in cedict:
            reading = cedict[char]["pinyin"]
            for d in cedict[char]["defs"]:
                norm = normalize_keyword(d)
                if norm not in used_keywords:
                    keyword = d
                    source = "cedict"
                    used_keywords.add(norm)
                    break

        if keyword is None and char in unihan:
            for d in unihan[char]:
                norm = normalize_keyword(d)
                if norm not in used_keywords:
                    keyword = d
                    source = "unihan"
                    used_keywords.add(norm)
                    break

        if keyword is None:
            base_def = None
            if char in cedict and cedict[char]["defs"]:
                base_def = cedict[char]["defs"][0]
                reading = cedict[char]["pinyin"]
            elif char in unihan and unihan[char]:
                base_def = unihan[char][0]

            if base_def:
                for i in range(2, 100):
                    candidate = f"{base_def} ({i})"
                    norm = normalize_keyword(candidate)
                    if norm not in used_keywords:
                        keyword = candidate
                        source = "numbered"
                        used_keywords.add(norm)
                        break

        if keyword:
            entry["keyword"] = keyword
            entry["reading"] = reading
            entry["source"] = source
            entry["components_detail"] = get_component_detail(char, entry["ids"], heisig_data)
            taiwan_results.append(entry)
        else:
            taiwan_failed.append(char)

    # Count numbered keywords
    numbered_count = sum(1 for e in mainland_results + taiwan_results if e.get("source") == "numbered")

    print(f"\nResults:")
    print(f"  Mainland: {len(mainland_results)} success, {len(mainland_failed)} failed")
    print(f"  Taiwan: {len(taiwan_results)} success, {len(taiwan_failed)} failed")
    print(f"  Keywords with numbers: {numbered_count}")

    # Save results
    with open(ROOT / "data" / "generated_mainland.json", "w", encoding="utf-8") as f:
        json.dump(mainland_results, f, ensure_ascii=False, indent=2)

    with open(ROOT / "data" / "generated_taiwan.json", "w", encoding="utf-8") as f:
        json.dump(taiwan_results, f, ensure_ascii=False, indent=2)

    print(f"\nSaved to data/generated_mainland.json and data/generated_taiwan.json")

    # Show some examples of numbered keywords
    print(f"\nExamples of numbered keywords:")
    numbered = [e for e in mainland_results + taiwan_results if e.get("source") == "numbered"][:10]
    for e in numbered:
        print(f"  {e['char']}: {e['keyword']}")


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == "--regenerate":
        regenerate_all()
    else:
        # Test with the "gnaw" characters
        cedict = parse_cedict()
        print("Testing 'gnaw' characters:")
        for char in ["啃", "嘬", "龁", "囓", "齕"]:
            if char in cedict:
                print(f"  {char}: {cedict[char]['defs']}")
        print("\nRun with --regenerate to regenerate all data")
