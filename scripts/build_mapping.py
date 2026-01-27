"""Build a unified mapping from characters/components to Heisig primitive names.

Strategy:
1. Start with everything named in RSH XML (characters + primitives + aliases)
2. Map CJK radical variants to their parent characters (e.g. 讠-> 言)
3. Map traditional/simplified variants to each other
4. Decompose all Excel characters via IDS
5. Report what's still unmapped
"""
import json
import re
import unicodedata
import openpyxl

# Load RSH parsed data
with open("data/rsh_parsed.json", "r", encoding="utf-8") as f:
    rsh = json.load(f)

# Build keyword lookup: character -> primary Heisig name
# and alias lookup: character -> list of all names
heisig_name = {}  # char -> primary keyword
heisig_aliases = {}  # char -> [all names including aliases]

for entry in rsh["characters"] + rsh["primitives"]:
    char = entry["character"]
    kw = entry["keyword"]
    aliases = entry["primitive_aliases"]
    heisig_name[char] = kw
    heisig_aliases[char] = [kw] + aliases

# Also build reverse lookup: keyword -> character
keyword_to_char = {}
for entry in rsh["characters"] + rsh["primitives"]:
    keyword_to_char[entry["keyword"]] = entry["character"]
    for alias in entry["primitive_aliases"]:
        keyword_to_char[alias] = entry["character"]

print(f"Heisig named entries: {len(heisig_name)}")

# --- Step 2: CJK Compatibility / Radical mappings ---
# Unicode has a "CJK Radicals Supplement" block (2E80-2EFF) and
# "Kangxi Radicals" (2F00-2FDF) that map to CJK unified chars.
# We can use Unicode decomposition mappings.

radical_to_parent = {}

# Method 1: Unicode decomposition mapping
for cp in range(0x2E80, 0x2FE0):
    ch = chr(cp)
    decomp = unicodedata.decomposition(ch)
    if decomp:
        # Format like "2F00" or "compat 2F00"
        parts = decomp.split()
        # Get the last hex value
        hex_val = parts[-1]
        try:
            parent = chr(int(hex_val, 16))
            radical_to_parent[ch] = parent
        except ValueError:
            pass

# Method 2: Known manual mappings for common simplified radicals
# These are simplified forms that appear as components
manual_radical_map = {
    '讠': '言',  # speech
    '钅': '金',  # metal
    '饣': '食',  # food
    '纟': '糸',  # thread
    '贝': '貝',  # shell
    '车': '車',  # vehicle
    '见': '見',  # see
    '门': '門',  # gate
    '鱼': '魚',  # fish
    '马': '馬',  # horse
    '鸟': '鳥',  # bird
    '页': '頁',  # page
    '风': '風',  # wind
    '韦': '韋',  # leather
    '长': '長',  # long
    '齿': '齒',  # teeth
    '龙': '龍',  # dragon
    '龟': '龜',  # turtle
    '亻': '人',  # person (side)
    '氵': '水',  # water (side)
    '扌': '手',  # hand (side)
    '忄': '心',  # heart (side)
    '犭': '犬',  # dog (side)
    '礻': '示',  # spirit/show (side)
    '衤': '衣',  # clothing (side)
    '⺗': '心',  # heart variant
    '⺝': '月',  # moon/flesh variant
    '⺼': '月',  # meat/flesh radical -> month (same in heisig)
    '⺶': '羊',  # sheep variant
    '灬': '火',  # fire dots (but Heisig calls this "cooking fire" separately)
    '⺌': '小',  # small top
    '⺊': '卜',  # divination
    '⺀': '八',  # eight variant
    '⺄': '乙',  # second variant
    '⺆': '冂',  # hood variant
    '⺈': '刀',  # knife variant
}

# Apply radical->parent, then check if parent has a Heisig name
mapped_via_radical = {}
for rad, parent in {**radical_to_parent, **manual_radical_map}.items():
    if rad not in heisig_name and parent in heisig_name:
        mapped_via_radical[rad] = (parent, heisig_name[parent])

print(f"Radicals mapped to Heisig via parent char: {len(mapped_via_radical)}")

# --- Step 3: Trad/Simplified cross-mapping from Excel ---
wb = openpyxl.load_workbook("data/Heisig's Remembering the Kanji vs. Hanzi v27.xlsx")
ws = wb["RTH+RSH+RTK"]

trad_to_simp = {}
simp_to_trad = {}
kanji_to_simp = {}

for row in ws.iter_rows(min_row=2, values_only=True):
    th, sh, k = row[3], row[4], row[5]
    if th and sh and th != sh:
        trad_to_simp[th] = sh
        simp_to_trad[sh] = th
    if k and sh and k != sh:
        kanji_to_simp[k] = sh

print(f"Trad->Simp pairs: {len(trad_to_simp)}")
print(f"Kanji->Simp pairs (where different): {len(kanji_to_simp)}")

# For trad/kanji chars not in Heisig, try mapping through simplified
mapped_via_variant = {}
for char_map in [trad_to_simp, kanji_to_simp]:
    for char, simp in char_map.items():
        if char not in heisig_name and simp in heisig_name:
            mapped_via_variant[char] = (simp, heisig_name[simp])

print(f"Trad/Kanji chars mapped to Heisig via simplified: {len(mapped_via_variant)}")

# --- Step 4: Parse IDS for all Excel characters ---
ids_map = {}
with open("data/IDS.TXT", "r", encoding="utf-8-sig") as f:
    for line in f:
        if line.startswith("#") or line.strip() == "":
            continue
        parts = line.strip().split("\t")
        if len(parts) >= 3:
            ids_map[parts[1]] = parts[2:]

IDS_OPERATORS = set("⿰⿱⿲⿳⿴⿵⿶⿷⿸⿹⿺⿻⿼⿽⿾⿿〾")

def parse_ids(ids_str):
    """Parse an IDS string into a tree structure.
    Returns (operator, [children]) or just the character."""
    cleaned = re.sub(r'\$\([^)]*\)', '', ids_str)
    cleaned = cleaned.replace('^', '').strip()

    tokens = list(cleaned)
    pos = 0

    def parse_next():
        nonlocal pos
        if pos >= len(tokens):
            return None
        ch = tokens[pos]
        pos += 1

        if ch in IDS_OPERATORS:
            # Determine number of children
            if ch in '⿰⿱⿴⿵⿶⿷⿸⿹⿺⿻⿼⿽⿾⿿':
                n_children = 2
            elif ch in '⿲⿳':
                n_children = 3
            elif ch == '〾':
                n_children = 1
            else:
                n_children = 2

            children = []
            for _ in range(n_children):
                child = parse_next()
                if child is not None:
                    children.append(child)
            return (ch, children)
        elif ch in '{}()0123456789 \t\r\n':
            # Skip formatting chars, try next
            return parse_next()
        else:
            return ch

    return parse_next()

def get_leaf_components(tree):
    """Get all leaf characters from a parsed IDS tree."""
    if tree is None:
        return []
    if isinstance(tree, str):
        return [tree]
    op, children = tree
    result = []
    for child in children:
        result.extend(get_leaf_components(child))
    return result

# --- Step 5: Build unified lookup ---
# Combine all naming sources
unified = {}

# Primary: direct Heisig names
for char, name in heisig_name.items():
    unified[char] = {
        "name": name,
        "aliases": heisig_aliases.get(char, []),
        "source": "heisig_direct",
    }

# Secondary: radical -> parent mapping
for rad, (parent, name) in mapped_via_radical.items():
    unified[rad] = {
        "name": name,
        "aliases": heisig_aliases.get(parent, []),
        "source": f"radical_of:{parent}",
    }

# Tertiary: trad/kanji -> simplified mapping
for char, (simp, name) in mapped_via_variant.items():
    if char not in unified:
        unified[char] = {
            "name": name,
            "aliases": heisig_aliases.get(simp, []),
            "source": f"variant_of:{simp}",
        }

print(f"\nUnified lookup size: {len(unified)}")

# --- Step 6: Decompose all Excel chars and check coverage ---
all_excel = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    for col in [3, 4, 5]:
        if row[col]:
            all_excel.add(row[col])

fully_mapped = 0  # char is in unified AND all its IDS components are in unified
partially_mapped = 0
unmapped_chars = 0
unmapped_components = set()
unmapped_detail = []

for char in sorted(all_excel):
    char_known = char in unified

    if char in ids_map:
        tree = parse_ids(ids_map[char][0])
        leaves = get_leaf_components(tree)
        unknown_leaves = [l for l in leaves if l not in unified and l != char]

        if char_known and not unknown_leaves:
            fully_mapped += 1
        elif char_known or not unknown_leaves:
            partially_mapped += 1
        else:
            unmapped_chars += 1

        for l in unknown_leaves:
            unmapped_components.add(l)
    else:
        if char_known:
            fully_mapped += 1
        else:
            unmapped_chars += 1

print(f"\nDecomposition coverage for {len(all_excel)} Excel characters:")
print(f"  Fully mapped (char + all components named):  {fully_mapped}")
print(f"  Partially mapped:                            {partially_mapped}")
print(f"  Unmapped:                                    {unmapped_chars}")
print(f"\nUnique unmapped components: {len(unmapped_components)}")

# Categorize unmapped components
strokes = set()
radicals = set()
cjk_chars = set()
other = set()

for c in unmapped_components:
    cp = ord(c)
    if 0x31C0 <= cp <= 0x31EF:
        strokes.add(c)
    elif 0x2E80 <= cp <= 0x2FDF:
        radicals.add(c)
    elif (0x4E00 <= cp <= 0x9FFF or 0x3400 <= cp <= 0x4DBF or
          0x20000 <= cp <= 0x2FA1F):
        cjk_chars.add(c)
    else:
        other.add(c)

print(f"  Strokes:            {len(strokes)}")
print(f"  Radical forms:      {len(radicals)}")
print(f"  CJK characters:     {len(cjk_chars)}")
print(f"  Other:              {len(other)}")

# Show the most frequently used unmapped components
from collections import Counter
comp_freq = Counter()
for char in all_excel:
    if char in ids_map:
        tree = parse_ids(ids_map[char][0])
        leaves = get_leaf_components(tree)
        for l in leaves:
            if l in unmapped_components:
                comp_freq[l] += 1

print(f"\nTop 30 most-used unmapped components:")
for comp, count in comp_freq.most_common(30):
    cp = ord(comp)
    name = unicodedata.name(comp, "?")
    print(f"  {comp} (U+{cp:04X}) used {count}x - {name}")

# Save unified mapping
with open("data/unified_mapping.json", "w", encoding="utf-8") as f:
    json.dump(unified, f, ensure_ascii=False, indent=2)

# Save unmapped components for manual review
unmapped_for_review = []
for comp, count in comp_freq.most_common():
    cp = ord(comp)
    name = unicodedata.name(comp, "?")
    ids_decomp = ids_map.get(comp, ["?"])[0] if comp in ids_map else "?"
    unmapped_for_review.append({
        "character": comp,
        "codepoint": f"U+{cp:04X}",
        "unicode_name": name,
        "frequency": count,
        "ids": ids_decomp,
    })

with open("data/unmapped_components.json", "w", encoding="utf-8") as f:
    json.dump(unmapped_for_review, f, ensure_ascii=False, indent=2)

print(f"\nSaved: data/unified_mapping.json ({len(unified)} entries)")
print(f"Saved: data/unmapped_components.json ({len(unmapped_for_review)} entries for review)")
